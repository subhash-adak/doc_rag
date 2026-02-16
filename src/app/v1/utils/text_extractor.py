import PyPDF2
from docx import Document
import openpyxl
import xlrd
from typing import Tuple, Dict

class TextExtractor:
    
    @staticmethod
    def extract_from_pdf(file_path: str) -> Tuple[str, Dict]:
        """Extract text from PDF"""
        text_parts = []
        metadata = {"page_count": 0}
        
        try:
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                metadata["page_count"] = len(reader.pages)
                
                for page_num, page in enumerate(reader.pages, 1):
                    page_text = page.extract_text()
                    if page_text.strip():
                        text_parts.append(f"\n--- Page {page_num} ---\n{page_text}")
            
            text = "\n".join(text_parts)
            return text.strip(), metadata
        
        except Exception as e:
            raise Exception(f"Error extracting PDF: {str(e)}")
    
    @staticmethod
    def extract_from_docx(file_path: str) -> Tuple[str, Dict]:
        """Extract text from DOCX"""
        try:
            doc = Document(file_path)
            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            text = "\n\n".join(paragraphs)
            
            metadata = {
                "paragraph_count": len(paragraphs),
                "page_count": None
            }
            
            return text.strip(), metadata
        
        except Exception as e:
            raise Exception(f"Error extracting DOCX: {str(e)}")
    
    @staticmethod
    def extract_from_txt(file_path: str) -> Tuple[str, Dict]:
        """Extract text from TXT"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                text = file.read()
            
            metadata = {
                "line_count": text.count('\n') + 1,
                "page_count": None
            }
            
            return text.strip(), metadata
        
        except Exception as e:
            raise Exception(f"Error extracting TXT: {str(e)}")
    
    @staticmethod
    def extract_from_xlsx(file_path: str) -> Tuple[str, Dict]:
        """Extract text from XLSX/XLS"""
        text_parts = []
        sheet_count = 0
        
        try:
            if file_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheet_count = len(wb.sheetnames)
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                    
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        if row_text.strip():
                            text_parts.append(row_text)
            
            else:  # .xls
                wb = xlrd.open_workbook(file_path)
                sheet_count = wb.nsheets
                
                for sheet in wb.sheets():
                    text_parts.append(f"\n=== Sheet: {sheet.name} ===\n")
                    
                    for row_idx in range(sheet.nrows):
                        row_text = " | ".join([str(cell.value) for cell in sheet.row(row_idx)])
                        if row_text.strip():
                            text_parts.append(row_text)
            
            text = "\n".join(text_parts)
            metadata = {
                "sheet_count": sheet_count,
                "page_count": None
            }
            
            return text.strip(), metadata
        
        except Exception as e:
            raise Exception(f"Error extracting Excel: {str(e)}")
    
    @staticmethod
    def extract_text(file_path: str, file_type: str) -> Tuple[str, Dict]:
        """Main extraction method"""
        extractors = {
            'PDF': TextExtractor.extract_from_pdf,
            'DOCX': TextExtractor.extract_from_docx,
            'TXT': TextExtractor.extract_from_txt,
            'XLSX': TextExtractor.extract_from_xlsx,
            'XLS': TextExtractor.extract_from_xlsx,
        }
        
        extractor = extractors.get(file_type.upper())
        if not extractor:
            raise ValueError(f"Unsupported file type: {file_type}")
        
        return extractor(file_path)